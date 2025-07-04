"""
pareto_sort.py

Loads an Excel workbook, performs non-dominated sorting on 4 objectives (columns F:I) for rows 3-452,
writes Pareto front numbers to column R for each row, and saves the workbook.
"""

import numpy as np
from openpyxl import load_workbook


def non_dominated_sort(scores):
    """
    Perform non-dominated sorting on a set of points.
    scores: numpy array of shape (n_points, n_objectives)
    Returns: numpy array of front numbers (1 = best Pareto front).
    """
    n_points = scores.shape[0]
    # Count of how many points dominate each point
    domination_count = np.zeros(n_points, dtype=int)
    # For each point, list of points it dominates
    dominated_sets = [set() for _ in range(n_points)]
    # Array to store Pareto front number for each point
    front_numbers = np.zeros(n_points, dtype=int)
    current_front = []

    # Compare each pair of points
    for p in range(n_points):
        for q in range(n_points):
            if p == q:
                continue
            # If p dominates q
            if np.all(scores[p] <= scores[q]) and np.any(scores[p] < scores[q]):
                dominated_sets[p].add(q)
            # If q dominates p
            elif np.all(scores[q] <= scores[p]) and np.any(scores[q] < scores[p]):
                domination_count[p] += 1
        # If no one dominates p, it's on the first front
        if domination_count[p] == 0:
            front_numbers[p] = 1
            current_front.append(p)

    front = 1
    # Extract subsequent fronts
    while current_front:
        next_front = []
        for p in current_front:
            for q in dominated_sets[p]:
                domination_count[q] -= 1
                if domination_count[q] == 0:
                    front_numbers[q] = front + 1
                    next_front.append(q)
        front += 1
        current_front = next_front

    return front_numbers


def main():
    files = ["Combined_results_DS_1_0.15.xlsx",
             "Combined_results_DS_1_0.30.xlsx",
             "Combined_results_DS_1_0.45.xlsx",
             "Combined_results_DS_10_0.15.xlsx",
             "Combined_results_DS_10_0.30.xlsx",
             "Combined_results_DS_10_0.45.xlsx",
             "Combined_results_SS_1_0.15.xlsx",
             "Combined_results_SS_1_0.30.xlsx",
             "Combined_results_SS_1_0.45.xlsx",
             "Combined_results_SS_10_0.15.xlsx",
             "Combined_results_SS_10_0.30.xlsx",
             "Combined_results_SS_10_0.45.xlsx",
             ]
    for file in files:
        # UNC path to the Excel file
        file_path = fr"Placeholder"

        # Load workbook and select the active sheet
        wb = load_workbook(filename=file_path)
        ws = wb.active

        # Read objectives from columns F (6) to I (9), rows 3 to 452
        scores = []
        for row in ws.iter_rows(min_row=3, max_row=452, min_col=6, max_col=9, values_only=True):
            scores.append(row)
        scores = np.array(scores, dtype=float)

        # Compute Pareto fronts
        fronts = non_dominated_sort(scores)

        # Write Pareto front numbers to column R (18), starting at row 3
        for idx, front in enumerate(fronts, start=3):
            ws.cell(row=idx, column=18, value=int(front))

        # Save the workbook (overwrites original)
        wb.save(filename=file_path)
        print(f"Pareto fronts written to column R and workbook saved: {file_path}")


if __name__ == '__main__':
    main()
